import winsound

def text(top_p,temperature, presence_penalty,frequency_penalty):
    # -*- coding: utf-8 -*-
    import json
    import time
    import requests
    with open("answer.json", "w", encoding="utf-8"):
        i =1
    # Function to read JSON file

    def readJson(json_name):
        with open(json_name, 'r', encoding='utf-8') as file:
            data = json.load(file)

        myQuestion = ["qwq"]  # 0 index as placeholder
        tiaoshi_num = 1
        for item in data:
            name = item.get('instruction', '默认值')
            myQuestion.append(name)
            print(f"读取json，读到第{tiaoshi_num}个了")
            tiaoshi_num += 1
        return myQuestion

    """This is the storage location of the key."""
    api = ["hk-xwmrry1000057382a22c07aef0d839d6de1793e49c17e3f6",
           ]
    key_num = 0
    model = "gpt-4o-mini"  # Changed to use OpenAI-HK model
    QuestionNumberStart = 1
    yourQuestionNumber = 10
    answer_filename = "answer.json"
    json_name = "wenti.json"

    # Prompt settings
    oneMyPromptSentence = ""
    twoMyPromptSentence = ""

    # ========== API 配置 ==========
    api_key = api[key_num]
    api_url = "https://api.openai-hk.com/v1/chat/completions"

    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {api_key}"
    }

    # ==================
    # call_model 函数开始 (modified for OpenAI-HK API)
    # ==================
    def call_model(prompt, retries=3, key_num=key_num):
        """Call the OpenAI-HK API to get the model response"""
        for attempt in range(retries):
            try:
                data = {
                    "max_tokens": 500,
                    "model": model,
                    "temperature": temperature,
                    "top_p": top_p,
                    "presence_penalty": presence_penalty,
                    "frequency_penalty": frequency_penalty,
                    "messages": [
                        {
                            "role": "system",
                            "content": ""
                        },
                        {
                            "role": "user",
                            "content": f"{oneMyPromptSentence}{prompt}{twoMyPromptSentence}"
                        }
                    ]
                }

                response = requests.post(api_url, headers=headers, data=json.dumps(data).encode('utf-8'))
                response.raise_for_status()  # Raise exception for HTTP errors

                result = response.content.decode("utf-8")
                result_json = json.loads(result)
                res_final = result_json["choices"][0]["message"]["content"]

                return res_final, ""

            except requests.exceptions.HTTPError as e:
                if response.status_code == 429:
                    wait_time = 10 + attempt * 5
                    print(f"⚠️ 请求过快 (429)，第 {attempt + 1} 次重试中... 等待 {wait_time} 秒")
                    time.sleep(wait_time)
                else:
                    print(f"API call error: {e}")
                    return "ERROR", ""
            except Exception as e:
                print(f"Other errors: {e}")
                return "ERROR", ""
        return "ERROR", ""


    question_WillInput = readJson(json_name)

    # Prepare output file
    file_willWrite = open(answer_filename, "w", encoding="utf-8")
    file_willWrite.write("[\n")
    answer_dic = {}

    while QuestionNumberStart <= yourQuestionNumber:
        base_response = call_model(question_WillInput[QuestionNumberStart])
        print(f"【{QuestionNumberStart}】Real-time output of program status - the answer provided by the model.: {base_response[0]}")
        temp_dict = {"output": base_response[0]}
        file_willWrite.write(json.dumps(temp_dict, indent=4, ensure_ascii=False))

        if QuestionNumberStart != yourQuestionNumber:
            file_willWrite.write(",")
        file_willWrite.write("\n")

        time.sleep(2)
        QuestionNumberStart += 1

    # File writing completed.
    file_willWrite.write("]\n")
    file_willWrite.close()
    print(f"The model has finished answering automatically, and the answers have been stored in a file.【{answer_filename}】中")


def data():
    import json
    from bert_score import BERTScorer

    import nltk
    # nltk.download('wordnet')
    from nltk.translate.bleu_score import sentence_bleu, SmoothingFunction
    from nltk.translate.meteor_score import meteor_score

    from sklearn.feature_extraction.text import TfidfVectorizer
    from sklearn.metrics.pairwise import cosine_similarity

    import torch
    from transformers import GPT2LMHeadModel, GPT2Tokenizer

    import hashlib
    import math

    # =====================参数修改================================
    my_answer_filename = "answer.json"  # The filename of the answer to be evaluated
    my_json_filename = "wenti.json"  # The JSON file name used for comparison evaluation
    my_num = 10  # The amount of data is used to calculate the average indicator.



    class SimHash:
        def __init__(self, tokens, hash_bits=64):
            self.hash_bits = hash_bits
            self.hash = self._calculate_hash(tokens)

        def _calculate_hash(self, tokens):
            # Initialize a vector of length hash_bits.
            v = [0] * self.hash_bits

            for token in tokens:
                # Calculate the hash value for each token
                hash_value = self._hash(token)
                # Convert the hash value to binary representation
                binary_hash = bin(hash_value)[2:].rjust(self.hash_bits, '0')

                # Update vector
                for i in range(self.hash_bits):
                    if binary_hash[i] == '1':
                        v[i] += 1
                    else:
                        v[i] -= 1

            # Convert the vector into the final hash value.
            fingerprint = 0
            for i in range(self.hash_bits):
                if v[i] > 0:
                    fingerprint |= (1 << i)

            return fingerprint

        def _hash(self, token):
            # Use hashlib to calculate the hash value.
            return int(hashlib.md5(token.encode('utf-8')).hexdigest(), 16)

        def distance(self, other):
            # Calculate the Hamming distance
            x = (self.hash ^ other.hash) & ((1 << self.hash_bits) - 1)
            return bin(x).count('1')

    # Load the pre-trained GPT-2 model and tokenizer.
    perplexitymodel = GPT2LMHeadModel.from_pretrained('gpt2')
    perplexitytokenizer = GPT2Tokenizer.from_pretrained('gpt2')

    def compute_perplexity(text):
        inputs = perplexitytokenizer.encode(text, return_tensors="pt")
        outputs = perplexitymodel(inputs, labels=inputs)
        loss = outputs.loss
        perplexity = torch.exp(loss)
        return perplexity.item()


    scorer = BERTScorer(
        lang="en",
        model_type="albert-large-v2",
        num_layers=5,
        rescale_with_baseline=True,
        device="cpu",
    )
    print("开始评判")

    with open(my_json_filename, 'r', encoding='utf-8') as file:
        data = json.load(file)


    array1 = []
    i = 0
    for item in data:
        name = item.get('output', '默认值')  # Use the get method to avoid KeyError.
        array1.append(name)
        i += 1
        # print(name)
    i = 0

    precision = 0
    recall = 0
    f1 = 0
    bleu = 0
    Cosine = 0
    perplexity = 0
    Hamming = 0
    METEOR = 0
    num = my_num

    # Use TF-IDF vectors to compute cosine similarity.
    def tfidf_cosine_similarity(text1, text2):
        # Initialize TF-IDF vectorizer
        vectorizer = TfidfVectorizer()
        # Convert the text into TF-IDF vectors.
        tfidf_matrix = vectorizer.fit_transform([text1, text2])
        # Calculate cosine similarity
        similarity = cosine_similarity(tfidf_matrix[0:1], tfidf_matrix[1:2])
        return similarity[0][0]



    with open(my_answer_filename, 'r', encoding='utf-8') as f:
        answer_data = json.load(f)

        for item in answer_data:
            line = item.get('output', '默认值')
            p = array1[i]
            i += 1


            references = [p]
            candidates = [line]
            P, R, F1 = scorer.score(candidates, references)
            print(f"Precision: {P.mean():.4f}, Recall: {R.mean():.4f}, F1: {F1.mean():.4f}")
            precision += P.mean()
            recall += R.mean()
            f1 += F1.mean()
            print()

            # Calculate METEOR score
            metorreference = [p.split()]
            metorcandidate = line.split()
            score = meteor_score(metorreference, metorcandidate)
            print("METEOR Score:", score)
            METEOR += score

            tfidf_similarity = tfidf_cosine_similarity(p, line)
            print("TF-IDF Cosine Similarity:", tfidf_similarity)
            Cosine += tfidf_similarity
            # Split the string into a list of words
            reference1 = [p.split()]
            candidate1 = line.split()
            weights = [1, 0, 0, 0]
            # Use a smooth function
            smoothie = SmoothingFunction().method1
            # Calculate the BLEU score using smoothing.
            bleu_score = sentence_bleu(reference1, candidate1, weights=weights, smoothing_function=smoothie)
            print(f"BLEU Score: {bleu_score:.4f}")
            bleu += bleu_score

            text = line
            perplexity1 = compute_perplexity(text)
            print(f"Perplexity: {perplexity1}")
            perplexity += perplexity1

            tokens1 = p.split()
            tokens2 = line.split()
            simhash1 = SimHash(tokens1)
            simhash2 = SimHash(tokens2)
            hamming_distance = simhash1.distance(simhash2)
            print(f"hamming_distance: {hamming_distance}")
            Hamming += hamming_distance
        return ["{:.4f}".format(precision/num), "{:.4f}".format(recall/num), "{:.4f}".format(f1/num), "{:.4f}".format(bleu/num), "{:.4f}".format(Cosine/num), "{:.4f}".format(perplexity/num), "{:.4f}".format(Hamming/num), "{:.4f}".format(METEOR/num)]

def appendExcle(data):
    import openpyxl as op
    from openpyxl import load_workbook
    wb = load_workbook("测试.xlsx")
    ws = wb.worksheets[0]
    ws.append(data)
    wb.save("测试.xlsx")
def createExcel():
    import openpyxl as op
    wb = op.Workbook()
    ws = wb['Sheet']
    ws.append(title)
    wb.save("测试.xlsx")
def deletelast():
    import openpyxl as op
    from openpyxl import load_workbook
    wb = load_workbook("测试.xlsx")
    ws = wb.worksheets[0]
    ws.delete_rows(ws.max_row)
    ws.save("测试.xlsx")
"""四个参数的变化大小(必须用小数)"""
top_p_change = 0.2
temperature_change = 0.4
present_penalty_change = 1.0
frequency_penalty_change = 1.0
title = ['top_p', 'temperature', 'presence_penalty', 'frequency_penalty','precision','recall','bleu','Cosine','perplexity','Hamming','METEOR']
import pandas as pd

file_path = '测试.xlsx'  # Replace with your Excel file path

df = pd.read_excel(file_path)
try:
    f = open("测试.xlsx")
    f.close()
except FileNotFoundError:
    #If there is no such file, create one.
    createExcel()
if df.empty:
    #If there is no previous data in the document, start from the minimum value.
    top_p = 0
    temperature = 1
    present_penalty = -2.0
    frequency_penalty = -2.0
else:
    #Continue from the last value.
    last_row = df[['top_p', 'temperature', 'presence_penalty', 'frequency_penalty']].iloc[-1].tolist()
    top_p = last_row[0]
    temperature = last_row[1]
    present_penalty = last_row[2]
    frequency_penalty = last_row[3] + frequency_penalty_change
    if frequency_penalty > 2.0:
        frequency_penalty = -2.0
print(f"The four extracted attributes are:{top_p, temperature, present_penalty,frequency_penalty}")


import numpy as np
while top_p <= 1.0:
    while temperature <= 2.0:
        while present_penalty <=2.0:
            while frequency_penalty <=2.0:
                text(top_p, temperature, present_penalty, frequency_penalty)
                list = [str("{:.1f}".format(top_p)), str("{:.1f}".format(temperature)), str(present_penalty),str(frequency_penalty)]
                for i in data():
                    list.append(i)
                print(list)
                appendExcle(list)
                print(
                    f"The judgment is complete, the four coefficients are respectively：top_p:{top_p},temperature:{temperature},present_penalty:{present_penalty},frequency_penalty:{frequency_penalty}")
                frequency_penalty += frequency_penalty_change
            present_penalty += present_penalty_change
            frequency_penalty = -2.0
        temperature  += temperature_change
        present_penalty = -2.0
    temperature = 0.0
    top_p += top_p_change
